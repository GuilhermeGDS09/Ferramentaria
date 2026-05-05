from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, os, math, tempfile
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
BASE_DIR = os.path.dirname(__file__)
DB = os.path.join(BASE_DIR, "database", "ferramentaria.db")

# ─────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────
def get_db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con

def init_db():
    os.makedirs(os.path.dirname(DB), exist_ok=True)
    con = get_db()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS maquinas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        tipo TEXT,
        turnos INTEGER DEFAULT 1,
        dias INTEGER DEFAULT 5,
        eficiencia REAL DEFAULT 85,
        parada REAL DEFAULT 2
    );

    CREATE TABLE IF NOT EXISTS pecas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT NOT NULL,
        descricao TEXT,
        material TEXT,
        setup_min REAL DEFAULT 30,
        lote_minimo INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS operacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        peca_id INTEGER NOT NULL,
        sequencia INTEGER,
        descricao TEXT,
        maquina_id INTEGER,
        tempo_min REAL,
        FOREIGN KEY (peca_id) REFERENCES pecas(id) ON DELETE CASCADE,
        FOREIGN KEY (maquina_id) REFERENCES maquinas(id)
    );

    CREATE TABLE IF NOT EXISTS ordens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero TEXT NOT NULL,
        peca_id INTEGER,
        quantidade INTEGER DEFAULT 1,
        data_inicio TEXT,
        prioridade TEXT DEFAULT 'Normal',
        cliente TEXT,
        status TEXT DEFAULT 'Aberta',
        FOREIGN KEY (peca_id) REFERENCES pecas(id)
    );
    """)
    con.commit()
    con.close()

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def cap_liquida(m):
    return max(0, ((m["turnos"] * 8 * m["dias"]) - m["parada"]) * (m["eficiencia"] / 100))

def calc_schedule(ordens, pecas_map, ops_map, maquinas_map):
    prio_val = {"Alta": 3, "Normal": 2, "Baixa": 1}
    sorted_orders = sorted(
        ordens,
        key=lambda o: (-prio_val.get(o["prioridade"], 2), o["data_inicio"] or "")
    )
    maq_end = {}
    results = []
    for o in sorted_orders:
        p = pecas_map.get(o["peca_id"])
        if not p:
            continue
        ops = sorted(ops_map.get(o["peca_id"], []), key=lambda x: x["sequencia"])
        try:
            cur = datetime.strptime(o["data_inicio"], "%Y-%m-%d").date()
        except Exception:
            cur = date.today()
        op_results = []
        for op in ops:
            m = maquinas_map.get(op["maquina_id"])
            if not m:
                continue
            cap = cap_liquida(m)
            cap_day = cap / m["dias"] if m["dias"] > 0 else 0
            h = (p["setup_min"] / 60) + (op["tempo_min"] / 60 * o["quantidade"])
            d_need = max(1, math.ceil(h / cap_day)) if cap_day > 0 else 999
            maq_free = maq_end.get(op["maquina_id"])
            start_op = max(cur, maq_free) if maq_free else cur
            end_op = start_op + timedelta(days=d_need)
            maq_end[op["maquina_id"]] = end_op
            cur = end_op
            op_results.append({
                "maquina": m["nome"],
                "maquina_id": op["maquina_id"],
                "descricao": op["descricao"],
                "horas": round(h, 2),
                "dias": d_need,
                "inicio": start_op.strftime("%d/%m/%Y"),
                "conclusao": end_op.strftime("%d/%m/%Y"),
                "conclusao_raw": end_op.isoformat()
            })
        results.append({
            "ordem": dict(o),
            "ops": op_results,
            "conclusao": cur.strftime("%d/%m/%Y"),
            "conclusao_raw": cur.isoformat()
        })
    return results

# ─────────────────────────────────────────────
# ROTAS HTML
# ─────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ─────────────────────────────────────────────
# API — MÁQUINAS
# ─────────────────────────────────────────────
@app.route("/api/maquinas", methods=["GET"])
def list_maquinas():
    con = get_db()
    rows = con.execute("SELECT * FROM maquinas ORDER BY id").fetchall()
    con.close()
    data = []
    for r in rows:
        m = dict(r)
        m["cap_liquida"] = round(cap_liquida(m), 1)
        data.append(m)
    return jsonify(data)

@app.route("/api/maquinas", methods=["POST"])
def add_maquina():
    d = request.json
    if not d or not d.get("nome"):
        return jsonify({"ok": False, "erro": "Nome obrigatório"}), 400
    con = get_db()
    try:
        con.execute(
            "INSERT INTO maquinas(nome,tipo,turnos,dias,eficiencia,parada) VALUES (?,?,?,?,?,?)",
            (d["nome"], d.get("tipo", ""), int(d.get("turnos", 1)),
             int(d.get("dias", 5)), float(d.get("eficiencia", 85)), float(d.get("parada", 2)))
        )
        con.commit()
    finally:
        con.close()
    return jsonify({"ok": True})

@app.route("/api/maquinas/<int:mid>", methods=["DELETE"])
def del_maquina(mid):
    con = get_db()
    con.execute("DELETE FROM maquinas WHERE id=?", (mid,))
    con.commit()
    con.close()
    return jsonify({"ok": True})

# ─────────────────────────────────────────────
# API — PEÇAS  ← CORRIGIDO
# ─────────────────────────────────────────────
@app.route("/api/pecas", methods=["GET"])
def list_pecas():
    con = get_db()
    pecas = [dict(p) for p in con.execute("SELECT * FROM pecas ORDER BY id").fetchall()]
    for p in pecas:
        ops = con.execute("""
            SELECT o.*, m.nome AS maquina_nome
            FROM operacoes o
            LEFT JOIN maquinas m ON m.id = o.maquina_id
            WHERE o.peca_id = ?
            ORDER BY o.sequencia
        """, (p["id"],)).fetchall()
        p["operacoes"] = [dict(o) for o in ops]
    con.close()
    return jsonify(pecas)

@app.route("/api/pecas", methods=["POST"])
def add_peca():
    d = request.json
    # Validações explícitas
    if not d:
        return jsonify({"ok": False, "erro": "Dados não enviados"}), 400
    if not d.get("codigo"):
        return jsonify({"ok": False, "erro": "Código obrigatório"}), 400
    if not d.get("descricao"):
        return jsonify({"ok": False, "erro": "Descrição obrigatória"}), 400
    operacoes = d.get("operacoes", [])
    if not operacoes:
        return jsonify({"ok": False, "erro": "Adicione pelo menos uma operação"}), 400

    con = get_db()
    try:
        # Insere a peça
        cur = con.execute(
            "INSERT INTO pecas(codigo,descricao,material,setup_min,lote_minimo) VALUES (?,?,?,?,?)",
            (
                d["codigo"].strip(),
                d["descricao"].strip(),
                d.get("material", "").strip(),
                float(d.get("setup_min", 30)),
                int(d.get("lote_minimo", 1))
            )
        )
        peca_id = cur.lastrowid

        # Insere as operações do roteiro
        for i, op in enumerate(operacoes):
            if not op.get("maquina_id") or not op.get("descricao") or not op.get("tempo_min"):
                con.close()
                return jsonify({"ok": False, "erro": f"Operação {i+1} incompleta (máquina, descrição e tempo são obrigatórios)"}), 400
            con.execute(
                "INSERT INTO operacoes(peca_id,sequencia,descricao,maquina_id,tempo_min) VALUES (?,?,?,?,?)",
                (peca_id, i + 1, op["descricao"].strip(),
                 int(op["maquina_id"]), float(op["tempo_min"]))
            )

        con.commit()
    except Exception as e:
        con.rollback()
        con.close()
        return jsonify({"ok": False, "erro": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True})

@app.route("/api/pecas/<int:pid>", methods=["DELETE"])
def del_peca(pid):
    con = get_db()
    con.execute("DELETE FROM pecas WHERE id=?", (pid,))
    con.commit()
    con.close()
    return jsonify({"ok": True})

# ─────────────────────────────────────────────
# API — ORDENS
# ─────────────────────────────────────────────
@app.route("/api/ordens", methods=["GET"])
def list_ordens():
    con = get_db()
    rows = con.execute("""
        SELECT o.*, p.codigo AS peca_codigo, p.descricao AS peca_desc
        FROM ordens o
        LEFT JOIN pecas p ON p.id = o.peca_id
        ORDER BY o.id
    """).fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/ordens", methods=["POST"])
def add_ordem():
    d = request.json
    if not d or not d.get("numero") or not d.get("peca_id") or not d.get("data_inicio"):
        return jsonify({"ok": False, "erro": "Número, peça e data de início são obrigatórios"}), 400
    con = get_db()
    try:
        con.execute(
            "INSERT INTO ordens(numero,peca_id,quantidade,data_inicio,prioridade,cliente,status) VALUES (?,?,?,?,?,?,'Aberta')",
            (d["numero"], int(d["peca_id"]), int(d.get("quantidade", 1)),
             d["data_inicio"], d.get("prioridade", "Normal"), d.get("cliente", ""))
        )
        con.commit()
    except Exception as e:
        con.close()
        return jsonify({"ok": False, "erro": str(e)}), 500
    finally:
        con.close()
    return jsonify({"ok": True})

@app.route("/api/ordens/<int:oid>", methods=["PATCH"])
def patch_ordem(oid):
    d = request.json
    con = get_db()
    if "status" in d:
        con.execute("UPDATE ordens SET status=? WHERE id=?", (d["status"], oid))
    con.commit()
    con.close()
    return jsonify({"ok": True})

@app.route("/api/ordens/<int:oid>", methods=["DELETE"])
def del_ordem(oid):
    con = get_db()
    con.execute("DELETE FROM ordens WHERE id=?", (oid,))
    con.commit()
    con.close()
    return jsonify({"ok": True})

# ─────────────────────────────────────────────
# API — PROGRAMAÇÃO
# ─────────────────────────────────────────────
@app.route("/api/programacao")
def programacao():
    con = get_db()
    ordens = [dict(o) for o in con.execute(
        "SELECT * FROM ordens WHERE status='Aberta' ORDER BY id"
    ).fetchall()]
    pecas_map = {r["id"]: dict(r) for r in con.execute("SELECT * FROM pecas").fetchall()}
    ops_raw = con.execute("SELECT * FROM operacoes ORDER BY sequencia").fetchall()
    ops_map = {}
    for op in ops_raw:
        ops_map.setdefault(op["peca_id"], []).append(dict(op))
    maquinas_map = {r["id"]: dict(r) for r in con.execute("SELECT * FROM maquinas").fetchall()}
    con.close()

    result = calc_schedule(ordens, pecas_map, ops_map, maquinas_map)
    return jsonify(result)

# ─────────────────────────────────────────────
# EXPORTAÇÃO EXCEL
# ─────────────────────────────────────────────
@app.route("/api/exportar/excel")
def export_excel():
    con = get_db()
    maquinas_rows = [dict(r) for r in con.execute("SELECT * FROM maquinas ORDER BY id").fetchall()]
    pecas_rows    = [dict(r) for r in con.execute("SELECT * FROM pecas ORDER BY id").fetchall()]
    ops_raw       = [dict(r) for r in con.execute("""
        SELECT o.*, m.nome AS maquina_nome
        FROM operacoes o LEFT JOIN maquinas m ON m.id=o.maquina_id
        ORDER BY o.peca_id, o.sequencia
    """).fetchall()]
    ordens_rows   = [dict(r) for r in con.execute("""
        SELECT o.*, p.codigo AS peca_codigo, p.descricao AS peca_desc
        FROM ordens o LEFT JOIN pecas p ON p.id=o.peca_id ORDER BY o.id
    """).fetchall()]
    pecas_map     = {r["id"]: r for r in pecas_rows}
    ops_map       = {}
    for op in ops_raw:
        ops_map.setdefault(op["peca_id"], []).append(op)
    maq_map = {r["id"]: r for r in maquinas_rows}
    ordens_abertas = [o for o in ordens_rows if o["status"] == "Aberta"]
    schedule = calc_schedule(ordens_abertas, pecas_map, ops_map, maq_map)
    con.close()

    # Estilos
    wb = Workbook()
    thin  = Side(style="thin", color="CBD5E1")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR = "1A3A5C"; HDR_F = "FFFFFF"; ALT = "EEF4FB"
    GN_B = "D6F0E3"; GN_F = "0F6E56"
    AM_B = "FEF3C7"; AM_F = "854F0B"
    RD_B = "FECACA"; RD_F = "A32D2D"

    def hf(color, bold=False, size=9):
        return Font(name="Arial", color=color, bold=bold, size=size)
    def hb(color): return PatternFill("solid", fgColor=color)
    def ct(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lf(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def hrow(ws, row, cols, texts):
        for col, t in zip(cols, texts):
            c = ws.cell(row=row, column=col, value=t)
            c.fill = hb(HDR); c.font = hf(HDR_F, True, 9)
            c.border = bdr; c.alignment = ct()

    def dc(ws, row, col, val, alt=False, align="left", bg=None, fg="000000", bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hb(bg if bg else (ALT if alt else "FFFFFF"))
        c.font = hf(fg, bold, 9); c.border = bdr
        c.alignment = ct() if align == "center" else lf()
        return c

    def title(ws, text, span):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(row=1, column=1, value=text)
        c.fill = hb("2563A8"); c.font = hf("FFFFFF", True, 12); c.alignment = ct()
        ws.row_dimensions[1].height = 28

    from openpyxl.utils import get_column_letter
    def cw(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    # Aba 1: Máquinas
    ws1 = wb.active; ws1.title = "Máquinas"
    title(ws1, "MÁQUINAS CADASTRADAS", 8)
    hrow(ws1, 3, range(1,9), ["Nome","Tipo","Turnos","Dias/sem","Eficiência","Parada(h)","Cap.bruta(h/sem)","Cap.líq.(h/sem)"])
    for i, m in enumerate(maquinas_rows):
        r = 4+i; alt = i%2==1
        cb = round(m["turnos"]*8*m["dias"] - m["parada"], 1)
        cl = round(cap_liquida(m), 1)
        dc(ws1,r,1,m["nome"],alt,bold=True); dc(ws1,r,2,m["tipo"],alt)
        dc(ws1,r,3,m["turnos"],alt,"center"); dc(ws1,r,4,m["dias"],alt,"center")
        dc(ws1,r,5,f'{m["eficiencia"]}%',alt,"center"); dc(ws1,r,6,m["parada"],alt,"center")
        dc(ws1,r,7,f"{cb}h",alt,"center")
        dc(ws1,r,8,f"{cl}h",alt,"center",bg=GN_B,fg=GN_F,bold=True)
    cw(ws1,{1:26,2:22,3:10,4:10,5:12,6:12,7:18,8:18})
    ws1.sheet_properties.tabColor = "1A3A5C"

    # Aba 2: Peças
    ws2 = wb.create_sheet("Peças e Roteiros")
    title(ws2, "PEÇAS E ROTEIROS DE FABRICAÇÃO", 7)
    hrow(ws2, 3, range(1,8), ["Código","Descrição","Material","Setup(min)","Lote mín.","Nº Op.","Roteiro"])
    r = 4
    for p in pecas_rows:
        alt = r%2==0
        ops_p = [o for o in ops_raw if o["peca_id"] == p["id"]]
        seq   = " → ".join(f'{o["sequencia"]}.{o["descricao"]} ({o["maquina_nome"]})' for o in ops_p)
        dc(ws2,r,1,p["codigo"],alt,bold=True); dc(ws2,r,2,p["descricao"],alt)
        dc(ws2,r,3,p["material"],alt); dc(ws2,r,4,p["setup_min"],alt,"center")
        dc(ws2,r,5,p["lote_minimo"],alt,"center"); dc(ws2,r,6,len(ops_p),alt,"center")
        dc(ws2,r,7,seq,alt); r+=1
    cw(ws2,{1:12,2:28,3:18,4:12,5:10,6:8,7:60})
    ws2.sheet_properties.tabColor = "2563A8"

    # Aba 3: Ordens
    ws3 = wb.create_sheet("Ordens de Produção")
    title(ws3, "ORDENS DE PRODUÇÃO", 9)
    hrow(ws3, 3, range(1,10), ["Nº Ordem","Peça","Descrição","Qtd","Início","Prioridade","Cliente","Status","Previsão"])
    sched_map = {s["ordem"]["numero"]: s for s in schedule}
    r = 4
    for o in ordens_rows:
        alt = r%2==0; s = sched_map.get(o["numero"])
        end = s["conclusao"] if s else "—"
        pb = {"Alta":RD_B,"Normal":AM_B,"Baixa":GN_B}.get(o["prioridade"],AM_B)
        pf = {"Alta":RD_F,"Normal":AM_F,"Baixa":GN_F}.get(o["prioridade"],AM_F)
        dc(ws3,r,1,o["numero"],alt,bold=True); dc(ws3,r,2,o.get("peca_codigo",""),alt,"center")
        dc(ws3,r,3,o.get("peca_desc",""),alt); dc(ws3,r,4,o["quantidade"],alt,"center")
        dc(ws3,r,5,o["data_inicio"],alt,"center")
        dc(ws3,r,6,o["prioridade"],alt,"center",bg=pb,fg=pf,bold=True)
        dc(ws3,r,7,o.get("cliente",""),alt); dc(ws3,r,8,o["status"],alt,"center")
        dc(ws3,r,9,end,alt,"center",fg=GN_F,bold=True); r+=1
    cw(ws3,{1:14,2:10,3:26,4:8,5:12,6:12,7:20,8:12,9:14})
    ws3.sheet_properties.tabColor = "854F0B"

    # Aba 4: Programação
    ws4 = wb.create_sheet("Programação")
    title(ws4, "PROGRAMAÇÃO DETALHADA", 9)
    hrow(ws4, 3, range(1,10), ["Nº Ordem","Peça","Prioridade","Operação","Máquina","Horas","Dias","Início","Conclusão"])
    r = 4
    for s in schedule:
        o = s["ordem"]
        pb = {"Alta":RD_B,"Normal":AM_B,"Baixa":GN_B}.get(o["prioridade"],AM_B)
        pf = {"Alta":RD_F,"Normal":AM_F,"Baixa":GN_F}.get(o["prioridade"],AM_F)
        for op in s["ops"]:
            alt = r%2==0
            dc(ws4,r,1,o["numero"],alt,bold=True)
            dc(ws4,r,2,o.get("peca_codigo",""),alt,"center")
            dc(ws4,r,3,o["prioridade"],alt,"center",bg=pb,fg=pf,bold=True)
            dc(ws4,r,4,op["descricao"],alt); dc(ws4,r,5,op["maquina"],alt)
            dc(ws4,r,6,op["horas"],alt,"center"); dc(ws4,r,7,op["dias"],alt,"center")
            dc(ws4,r,8,op["inicio"],alt,"center")
            dc(ws4,r,9,op["conclusao"],alt,"center",fg=GN_F,bold=True); r+=1
    cw(ws4,{1:14,2:10,3:12,4:28,5:22,6:10,7:8,8:14,9:14})
    ws4.sheet_properties.tabColor = "0F6E56"

    # Aba 5: Carga por Máquina
    ws5 = wb.create_sheet("Carga por Máquina")
    title(ws5, "CARGA DE TRABALHO POR MÁQUINA", 7)
    hrow(ws5, 3, range(1,8), ["Máquina","Cap.líq.(h/sem)","Horas alocadas","Carga(%)","Folga(h)","Status","Ordens"])
    maq_load = {m["id"]: 0.0 for m in maquinas_rows}
    maq_ords = {m["id"]: [] for m in maquinas_rows}
    for o in ordens_abertas:
        p = pecas_map.get(o["peca_id"])
        if not p: continue
        for op in ops_map.get(o["peca_id"], []):
            h = (p["setup_min"]/60) + (op["tempo_min"]/60 * o["quantidade"])
            maq_load[op["maquina_id"]] = maq_load.get(op["maquina_id"],0) + h
            ords = maq_ords.get(op["maquina_id"],[])
            if o["numero"] not in ords: ords.append(o["numero"])
            maq_ords[op["maquina_id"]] = ords
    r = 4
    for m in maquinas_rows:
        alt = r%2==0; cap = round(cap_liquida(m),1)
        load = round(maq_load.get(m["id"],0),1)
        pct  = round(load/cap*100,1) if cap>0 else 0
        folga = round(cap-load,1)
        st  = "Sobrecarregado" if pct>90 else "Atenção" if pct>75 else "OK"
        sb  = RD_B if pct>90 else AM_B if pct>75 else GN_B
        sf  = RD_F if pct>90 else AM_F if pct>75 else GN_F
        dc(ws5,r,1,m["nome"],alt,bold=True); dc(ws5,r,2,f"{cap}h",alt,"center")
        dc(ws5,r,3,f"{load}h",alt,"center")
        dc(ws5,r,4,f"{pct}%",alt,"center",bg=sb,fg=sf,bold=True)
        dc(ws5,r,5,f"{folga}h",alt,"center")
        dc(ws5,r,6,st,alt,"center",bg=sb,fg=sf,bold=True)
        dc(ws5,r,7,", ".join(maq_ords.get(m["id"],[])),alt); r+=1
    cw(ws5,{1:26,2:16,3:16,4:12,5:12,6:16,7:40})
    ws5.sheet_properties.tabColor = "A32D2D"

    path = os.path.join(tempfile.gettempdir(), f"ferramentaria_{date.today().isoformat()}.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True,
                     download_name=f"ferramentaria_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    print("\n✅  Ferramentaria rodando em: http://localhost:5000\n")
    app.run(debug=True, port=5000)

